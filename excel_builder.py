"""
excel_builder.py
Produces the content XLSX from content_generator output.

Column structure (A–F locked, G–J are additions for the new format):
  A  Lesson Number   1 / 2 / 3
  B  Lesson          Vocabulary / Retrieval / Inference
  C  Version         Standard / Supported
  D  Section         Text / Question / Answer
  E  Label           Q1–Q7 (or blank for text rows)
  F  Content         Question text or answer text
  G  Format          open_line / tick_one / true_false_table / etc.
  H  Marks           1 / 2 / 3
  I  Text Reference  The scoping instruction for the question
  J  Format Data     JSON string of format_data (options, items, statements etc.)
"""

import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colours (hex, no #)
# ---------------------------------------------------------------------------
DARK_BLUE  = "1A2C5E"
MID_BLUE   = "4472C4"
LIGHT_BLUE = "DCE6F1"
GREEN_FILL = "E2EFDA"
WHITE      = "FFFFFF"
GREY       = "F2F2F2"


def _header_fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)


def _thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def build_excel(content: dict, output_path: str) -> None:
    """
    Build the content XLSX from content_generator output.

    content      : dict returned by content_generator.generate_content()
    output_path  : full path to write the .xlsx file
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Being a Reader Content"

    # ── Column widths ──────────────────────────────────────────────────────
    col_widths = {
        "A": 8,    # Lesson Number
        "B": 14,   # Lesson
        "C": 12,   # Version
        "D": 12,   # Section
        "E": 8,    # Label
        "F": 60,   # Content
        "G": 22,   # Format
        "H": 7,    # Marks
        "I": 40,   # Text Reference
        "J": 50,   # Format Data (JSON)
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── Header row ─────────────────────────────────────────────────────────
    headers = [
        "Lesson Number", "Lesson", "Version", "Section",
        "Label", "Content", "Format", "Marks",
        "Text Reference", "Format Data",
    ]
    for col_i, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_i, value=header)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = _header_fill(DARK_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[1].height = 18

    row = 2
    lesson_num = 0

    for lesson in content.get("lessons", []):
        lesson_num += 1
        lesson_type = lesson.get("lesson_type", "").capitalize()
        questions = lesson.get("questions", [])
        text = lesson.get("text", "")  # may be empty if app didn't inject

        # Alternate lesson background
        lesson_fill = _header_fill(LIGHT_BLUE if lesson_num % 2 == 0 else GREY)

        def _write_row(version, section, label, content_text,
                       fmt="", marks="", ref="", fmt_data=""):
            nonlocal row
            values = [
                lesson_num,
                lesson_type,
                version,
                section,
                label,
                content_text,
                fmt,
                marks if marks != "" else None,
                ref,
                fmt_data,
            ]
            for col_i, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_i, value=val)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = _thin_border()
                cell.font = Font(size=9)
                # Shade lesson number + lesson name columns
                if col_i <= 2:
                    cell.fill = lesson_fill
            ws.row_dimensions[row].height = 15
            row += 1

        # ── Text row ──────────────────────────────────────────────────────
        if text:
            _write_row("Standard", "Text", "", text)

        # ── Question rows ─────────────────────────────────────────────────
        for q in questions:
            q_num = q.get("number", "")
            q_label = f"Q{q_num}"
            q_text = q.get("question", "")
            q_answer = q.get("answer", "")
            q_fmt = q.get("format", "")
            q_marks = q.get("marks", "")
            q_ref = q.get("text_reference", "")
            q_scaffold = q.get("supported_scaffold", "") or ""

            # Serialise format_data to compact JSON
            fd = q.get("format_data", {})
            fd_json = json.dumps(fd, ensure_ascii=False) if fd else ""

            # Standard question
            _write_row(
                "Standard", "Question", q_label,
                q_text, q_fmt, q_marks, q_ref, fd_json,
            )

            # Standard answer
            _write_row(
                "Standard", "Answer", q_label,
                q_answer, q_fmt, q_marks, q_ref, "",
            )

        # ── Supported question rows (Q1–Q5 only, with scaffold) ──────────
        for q in questions[:5]:
            q_num = q.get("number", "")
            q_label = f"Q{q_num}"
            q_text = q.get("question", "")
            q_answer = q.get("answer", "")
            q_fmt = q.get("format", "")
            q_marks = q.get("marks", "")
            q_ref = q.get("text_reference", "")
            q_scaffold = q.get("supported_scaffold", "") or ""

            fd = q.get("format_data", {})
            fd_json = json.dumps(fd, ensure_ascii=False) if fd else ""

            # Supported question (with scaffold appended if present)
            supported_q = q_text
            if q_scaffold:
                supported_q = f"{q_text}\n\nHint: {q_scaffold}"

            _write_row(
                "Supported", "Question", q_label,
                supported_q, q_fmt, q_marks, q_ref, fd_json,
            )

            # Supported answer
            _write_row(
                "Supported", "Answer", q_label,
                q_answer, q_fmt, q_marks, q_ref, "",
            )

        # ── I can statements ──────────────────────────────────────────────
        for ican in lesson.get("i_can_statements", []):
            _write_row("", "I can", "", ican)

    # ── Freeze top row ─────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:J{row - 1}"

    wb.save(output_path)
